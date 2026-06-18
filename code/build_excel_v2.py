import csv, json, openpyxl
from collections import defaultdict, Counter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Load data ─────────────────────────────────────────────────────
def load_csv(path):
    with open(path, newline='', encoding='utf-8') as f:
        return list(csv.DictReader(f))

def load_json(path):
    with open(path) as f:
        return json.load(f)

md_A  = load_csv('/home2/muskan.singh/results/adverse_weather_moondream_v2.csv')
md_C  = load_csv('/home2/muskan.singh/results/adverse_weather_moondream_noimage_v2.csv')
pg_A  = load_csv('/home2/muskan.singh/results/adverse_weather_paligemma_v2.csv')
pg_C  = load_csv('/home2/muskan.singh/results/adverse_weather_paligemma_noimage_v2.csv')
md_an = load_json('/home2/muskan.singh/results/aw_moondream_v2_analysis.json')
pg_an = load_json('/home2/muskan.singh/results/aw_paligemma_v2_analysis.json')

manifest = load_json('/home2/muskan.singh/aw_manifest_v2.json')
var_lookup = {}
for entry in manifest:
    for q in entry['questions']:
        if q.get('use_for')=='linguistic_variation':
            var_lookup[q['q_id']] = {'variation_type':q.get('variation_type',''),'source_q_id':q.get('source_q_id','')}

# ── Styles ────────────────────────────────────────────────────────
NAVY='0D1B2A'; TEAL='0E7C7B'; AMBER='E8A838'; RED='C0392B'
GREEN='27AE60'; LIGHT='EAF4FB'; WHITE='FFFFFF'; GREY='F2F2F2'
BLUE='2980B9'; PURPLE='7C3AED'

thin = Border(
    left=Side(style='thin',color='CCCCCC'), right=Side(style='thin',color='CCCCCC'),
    top=Side(style='thin',color='CCCCCC'),  bottom=Side(style='thin',color='CCCCCC'))

def hdr(ws, row, col, val, bg=NAVY, fg=WHITE, wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill('solid', fgColor=bg)
    c.font = Font(color=fg, bold=True, name='Calibri', size=10)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    return c

def cell(ws, row, col, val, bg=WHITE, fg='000000', bold=False, align='center'):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill('solid', fgColor=bg)
    c.font = Font(color=fg, bold=bold, name='Calibri', size=10)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border = thin
    return c

def title_row(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.fill = PatternFill('solid', fgColor=NAVY)
    c.font = Font(color=WHITE, bold=True, name='Calibri', size=13)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

def set_cols(ws, widths):
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w

def sig_color(v):
    if v is None: return GREY
    if v >= 50:   return GREEN
    if v >= 20:   return AMBER
    if v >= 5:    return 'E67E22'
    return RED

def acc_color(v):
    if v is None: return GREY
    if v >= 70:   return GREEN
    if v >= 50:   return AMBER
    return RED

action_colors = {'KEEP':GREEN,'INVESTIGATE':AMBER,'BASELINE ONLY':BLUE,'REDESIGN':RED}

def get_action(s):
    if s['signal']=='NONE' and s.get('unclear',0)/max(s.get('n',1),1)>0.4: return 'REDESIGN'
    if s['signal']=='NONE': return 'BASELINE ONLY'
    if s['signal']=='WEAK': return 'INVESTIGATE'
    return 'KEEP'

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════
# Helper: build questions sheet for any model
# ══════════════════════════════════════════════════════════════════
def build_questions_sheet(ws, model_name, summary, vgs_metric_note):
    title_row(ws, '%s | Per-Question Analysis | main_eval original images' % model_name, 17)
    ws.freeze_panes = 'A3'
    h = ['Q_ID','Type','Question','N','GT=yes','GT=no',
         'A_acc%','C_acc%','VGS%','VGS_yes%','VGS_no%',
         'yes_acc%','no_acc%','Bias_gap%','Unclear','Signal','Action']
    w = [12,12,44,6,7,6,8,8,10,10,10,9,9,10,8,10,15]
    set_cols(ws, w)
    for i,(hh,ww) in enumerate(zip(h,w),1): hdr(ws,2,i,hh,bg=TEAL)
    # VGS metric note
    ws.merge_cells(start_row=30, start_column=1, end_row=30, end_column=17)
    note = ws.cell(row=30, column=1, value=vgs_metric_note)
    note.fill = PatternFill('solid', fgColor='FFF3CD')
    note.font = Font(name='Calibri', size=9, italic=True, color='7C4700')
    note.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 22
    for ri, q in enumerate(sorted(summary, key=lambda x:x['q_id']), 3):
        action = get_action(q)
        row_bg = LIGHT if ri%2==0 else WHITE
        vals = [q['q_id'],q['q_type'],q['question'],q['n'],q['gt_yes'],q['gt_no'],
                q['a_acc'],q['c_acc'],q['vgs'],q['vgs_yes'],q['vgs_no'],
                q['yes_acc'],q['no_acc'],q['bias_gap'],q['unclear'],q['signal'],action]
        for ci,v in enumerate(vals,1):
            bg=row_bg; fg='000000'; bold=False
            if ci==16: bg=sig_color(q['vgs_yes'] if q['vgs_yes'] is not None else q['vgs']); fg=WHITE; bold=True
            elif ci==17: bg=action_colors.get(action,GREY); fg=WHITE; bold=True
            elif ci in (7,8):
                bg=acc_color(v) if v else row_bg
                fg=WHITE if v and (v>=70 or v<50) else '000000'
            elif ci in (9,10,11) and v is not None:
                bg=GREEN if v>40 else (AMBER if v>10 else (RED if v<-10 else row_bg))
                fg=WHITE if abs(v)>10 else '000000'
            elif ci==14 and v is not None:
                bg=RED if v>25 else (AMBER if v>10 else (GREEN if abs(v)<=10 else BLUE))
                fg=WHITE if abs(v)>10 else '000000'; bold=True
            cell(ws,ri,ci,v,bg=bg,fg=fg,bold=bold,align='left' if ci==3 else 'center')
        ws.row_dimensions[ri].height = 32

# ══════════════════════════════════════════════════════════════════
# Helper: weather sheet
# ══════════════════════════════════════════════════════════════════
def build_weather_sheet(ws, model_name, weather_data, main_rows):
    title_row(ws, '%s | Accuracy by Weather Condition' % model_name, 9)
    ws.freeze_panes = 'A3'
    h = ['Condition','N_images','Overall_acc%','GT=yes_acc%','GT=no_acc%','Bias_gap%','GT_yes_n','GT_no_n','Unclear%']
    w = [13,10,13,13,13,11,11,10,10]
    set_cols(ws,w)
    for i,(hh,ww) in enumerate(zip(h,w),1): hdr(ws,2,i,hh,bg=TEAL)
    for ri,wd in enumerate(weather_data,3):
        row_bg=LIGHT if ri%2==0 else WHITE
        gap = round(wd['yes_acc']-wd['no_acc'],1) if wd.get('yes_acc') and wd.get('no_acc') else None
        unc_rows=[r for r in main_rows if r['weather']==wd['weather']]
        unc=round(sum(r['extracted_ans']=='unclear' for r in unc_rows)/len(unc_rows)*100,1) if unc_rows else 0
        vals=[wd['weather'].capitalize(),wd['n'],wd['acc'],wd.get('yes_acc'),wd.get('no_acc'),gap,wd['gt_yes'],wd['gt_no'],unc]
        for ci,v in enumerate(vals,1):
            bg=row_bg; fg='000000'; bold=False
            if ci==3: bg=acc_color(v); fg=WHITE; bold=True
            elif ci==6 and v is not None:
                bg=RED if v>25 else (AMBER if v>10 else GREEN); fg=WHITE; bold=True
            cell(ws,ri,ci,v,bg=bg,fg=fg,bold=bold)
        ws.row_dimensions[ri].height = 22
    # Overall row
    all_yes=[r for r in main_rows if r['ground_truth']=='yes']
    all_no =[r for r in main_rows if r['ground_truth']=='no']
    ov=round(sum(r['correct']=='True' for r in main_rows)/len(main_rows)*100,1)
    ya=round(sum(r['correct']=='True' for r in all_yes)/len(all_yes)*100,1)
    na=round(sum(r['correct']=='True' for r in all_no)/len(all_no)*100,1)
    gap=round(ya-na,1)
    unc=round(sum(r['extracted_ans']=='unclear' for r in main_rows)/len(main_rows)*100,1)
    for ci,v in enumerate(['OVERALL',len(main_rows),ov,ya,na,gap,len(all_yes),len(all_no),unc],1):
        cell(ws,8,ci,v,bg=NAVY,fg=WHITE,bold=True)

# ══════════════════════════════════════════════════════════════════
# Helper: SCS sheet
# ══════════════════════════════════════════════════════════════════
def build_scs_sheet(ws, model_name, scs_data):
    title_row(ws, '%s | Scene-Change Sensitivity (SCS) | Clear -> Fog/Rain/Snow' % model_name, 10)
    ws.freeze_panes = 'A3'
    h=['Q_ID','Question','N_pairs','SCS_overall%','SCS_fog%','SCS_rain%','SCS_snow%','GT_flip%','CDR%','Notes']
    w=[12,44,8,13,10,10,10,11,10,32]
    set_cols(ws,w)
    for i,(hh,ww) in enumerate(zip(h,w),1): hdr(ws,2,i,hh,bg=TEAL)
    for ri,d in enumerate(scs_data,3):
        row_bg=LIGHT if ri%2==0 else WHITE
        scs=d.get('scs'); gt_flip=d.get('gt_flip'); cdr=d.get('cdr')
        note=''
        if scs is not None:
            if scs>=50:   note='High — model responds to weather change'
            elif scs>=25: note='Moderate sensitivity'
            elif gt_flip is not None and gt_flip<10: note='Low GT flip — weather does not change GT here'
            else:         note='Low — model ignores weather change'
        vals=[d['q_id'],d.get('question',''),d.get('n_pairs'),scs,
              d.get('scs_fog'),d.get('scs_rain'),d.get('scs_snow'),gt_flip,cdr,note]
        for ci,v in enumerate(vals,1):
            bg=row_bg; fg='000000'; bold=False
            if ci==4 and v is not None:
                bg=GREEN if v>=50 else (AMBER if v>=25 else RED); fg=WHITE; bold=True
            elif ci in (5,6,7) and v is not None:
                bg=GREEN if v>=50 else (AMBER if v>=25 else RED); fg=WHITE
            cell(ws,ri,ci,v,bg=bg,fg=fg,bold=bold,align='left' if ci in (2,10) else 'center')
        ws.row_dimensions[ri].height = 30

# ══════════════════════════════════════════════════════════════════
# Helper: linguistic sheet
# ══════════════════════════════════════════════════════════════════
def build_ling_sheet(ws, model_name, ling_data):
    title_row(ws, '%s | Linguistic Variation Analysis | Paraphrase Consistency + Negation Accuracy' % model_name, 9)
    ws.freeze_panes = 'A3'
    h=['Source_Q_ID','Question','Para_pairs','Para_consistent%','Neg_pairs','Neg_acc%','Neg_flip%','Signal','Notes']
    w=[14,44,11,16,10,10,12,12,35]
    set_cols(ws,w)
    for i,(hh,ww) in enumerate(zip(h,w),1): hdr(ws,2,i,hh,bg=TEAL)
    sig_col={'ROBUST':GREEN,'MODERATE':AMBER,'INCONSISTENT':RED,'NEG FAILURE':RED,'N/A':GREY}
    for ri,d in enumerate(ling_data,3):
        row_bg=LIGHT if ri%2==0 else WHITE
        pp=d.get('para_pct'); na=d.get('neg_acc')
        note=''
        if pp is not None and pp<60:       note='Unstable — changes across paraphrases'
        elif na is not None and na<40:     note='Negation failure — model does not flip correctly'
        elif pp is not None and pp>=80 and na is not None and na>=60:
            note='Robust — consistent paraphrase + handles negation'
        vals=[d['source_q_id'],d.get('question',''),d.get('para_pairs'),pp,
              d.get('neg_pairs'),na,d.get('neg_flip'),d.get('signal'),note]
        for ci,v in enumerate(vals,1):
            bg=row_bg; fg='000000'; bold=False
            if ci==4 and v is not None:
                bg=GREEN if v>=80 else (AMBER if v>=60 else RED); fg=WHITE; bold=True
            elif ci==6 and v is not None:
                bg=GREEN if v>=60 else (AMBER if v>=40 else RED); fg=WHITE; bold=True
            elif ci==8:
                bg=sig_col.get(v,GREY); fg=WHITE if v!='N/A' else '000000'; bold=True
            cell(ws,ri,ci,v,bg=bg,fg=fg,bold=bold,align='left' if ci in (2,9) else 'center')
        ws.row_dimensions[ri].height = 32

# ══════════════════════════════════════════════════════════════════
# SHEET 1: README
# ══════════════════════════════════════════════════════════════════
ws = wb.active; ws.title = 'README'
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 85
title_row(ws, 'Adverse Weather Results Workbook | Can VLMs Understand Driving? | IIIT Hyderabad', 2)
readme = [
    ('PROJECT',      'Can VLMs Understand Driving? | IIIT Hyderabad | Supervisor: Shankar Gangisetty'),
    ('CATEGORY',     'Adverse Weather | BDD100K 108 images + 72 counterfactual | Manifest v2'),
    ('MODELS DONE',  'Moondream 2B | PaliGemma 3B'),
    ('MODELS RUNNING','SmolVLM 2.2B (job 2610436)'),
    ('MODELS PENDING','LLaVA-OV 8B | InternVL3 8B'),
    ('',''),
    ('VGS (ours)',    'Mode A accuracy - Mode C accuracy. NOTE: Sir flagged this needs rethinking towards IoU-like metric. Current VGS is informative but coarse.'),
    ('VGS_yes',      'VGS on GT=yes rows only. Clean signal for confounded models (Moondream defaults to no on blank image).'),
    ('SCS',          'Scene-Change Sensitivity = fraction of clear->augmented pairs where answer flipped.'),
    ('CDR',          'Correct Direction Rate = fraction of flipped pairs where flip was correct.'),
    ('Para_consistent','% of paraphrase variants where model gives same answer as original on same image.'),
    ('Neg_acc',      'Accuracy on negated variants. Low = negation failure (model flips but wrong).'),
    ('Bias_gap',     'GT=yes acc - GT=no acc. Positive = affirmation bias.'),
    ('Mode C note MD','Moondream defaults to ~100% no on blank images — VGS_yes is the valid metric, not VGS_overall.'),
    ('Mode C note PG','PaliGemma says unanswerable on blank images (reverted to unclear). VGS_overall valid since unanswerable is uniform across GT values.'),
    ('',''),
    ('SHEETS',''),
    ('MD_v2_Questions','Moondream per-question analysis'),
    ('MD_v2_Weather','Moondream weather breakdown'),
    ('MD_v2_SCS','Moondream counterfactual SCS'),
    ('MD_v2_Linguistic','Moondream linguistic variations'),
    ('PG_v2_Questions','PaliGemma per-question analysis'),
    ('PG_v2_Weather','PaliGemma weather breakdown'),
    ('PG_v2_SCS','PaliGemma counterfactual SCS'),
    ('PG_v2_Linguistic','PaliGemma linguistic variations'),
    ('CrossModel_VGS','VGS comparison across models — fills as models complete'),
    ('CrossModel_Bias','Affirmation bias gap comparison across models'),
    ('LangPrior_Analysis','Case study: when does model answer with/without image across models'),
    ('Question_Tracker','Final action per question across all models'),
]
for ri,(label,text) in enumerate(readme,2):
    bg=LIGHT if ri%2==0 else WHITE
    lc=ws.cell(row=ri,column=1,value=label)
    lc.fill=PatternFill('solid',fgColor=TEAL if label and label not in ('','SHEETS') else (AMBER if label=='SHEETS' else WHITE))
    lc.font=Font(color=WHITE if label else '000000',bold=bool(label),name='Calibri',size=10)
    lc.alignment=Alignment(horizontal='left',vertical='center')
    tc=ws.cell(row=ri,column=2,value=text)
    tc.fill=PatternFill('solid',fgColor=bg)
    tc.font=Font(name='Calibri',size=10)
    tc.alignment=Alignment(horizontal='left',vertical='center',wrap_text=True)
    ws.row_dimensions[ri].height=20

# ══════════════════════════════════════════════════════════════════
# MOONDREAM SHEETS
# ══════════════════════════════════════════════════════════════════
md_main = [r for r in md_A if r['use_for']=='main_eval' and r['is_augmented']=='False']

ws2 = wb.create_sheet('MD_v2_Questions')
build_questions_sheet(ws2, 'Moondream 2B v2', md_an['summary'],
    'VGS metric note: Moondream defaults to ~100% no on blank images. Use VGS_yes as clean signal, not VGS_overall.')

ws3 = wb.create_sheet('MD_v2_Weather')
build_weather_sheet(ws3, 'Moondream 2B v2', md_an['weather'], md_main)

ws4 = wb.create_sheet('MD_v2_SCS')
build_scs_sheet(ws4, 'Moondream 2B v2', md_an.get('scs', []))

ws5 = wb.create_sheet('MD_v2_Linguistic')
build_ling_sheet(ws5, 'Moondream 2B v2', md_an.get('linguistic', []))

# ══════════════════════════════════════════════════════════════════
# PALIGEMMA SHEETS
# ══════════════════════════════════════════════════════════════════
pg_main = [r for r in pg_A if r['use_for']=='main_eval' and r['is_augmented']=='False']

ws6 = wb.create_sheet('PG_v2_Questions')
build_questions_sheet(ws6, 'PaliGemma 3B v2', pg_an['summary'],
    'VGS metric note: PaliGemma says unanswerable on blank images (kept as unclear=incorrect). VGS_overall is valid — unanswerable is uniform across GT=yes/no.')

ws7 = wb.create_sheet('PG_v2_Weather')
build_weather_sheet(ws7, 'PaliGemma 3B v2', pg_an['weather'], pg_main)

ws8 = wb.create_sheet('PG_v2_SCS')
build_scs_sheet(ws8, 'PaliGemma 3B v2', pg_an['scs'])

ws9 = wb.create_sheet('PG_v2_Linguistic')
build_ling_sheet(ws9, 'PaliGemma 3B v2', pg_an['linguistic'])

# ══════════════════════════════════════════════════════════════════
# SHEET: CrossModel_VGS
# ══════════════════════════════════════════════════════════════════
ws10 = wb.create_sheet('CrossModel_VGS')
title_row(ws10, 'Cross-Model VGS Comparison | Adverse Weather | 2/5 models complete', 10)
ws10.freeze_panes = 'A3'
h=['Q_ID','Type','Question','MD_VGS','MD_signal','PG_VGS','PG_signal','SV_VGS','LV_VGS','IV_VGS']
w=[12,11,44,10,10,10,10,10,10,10]
set_cols(ws10,w)
for i,(hh,ww) in enumerate(zip(h,w),1): hdr(ws10,2,i,hh,bg=TEAL)

md_by_q = {s['q_id']:s for s in md_an['summary']}
pg_by_q = {s['q_id']:s for s in pg_an['summary']}
all_qids = sorted(md_by_q.keys())

for ri,q_id in enumerate(all_qids,3):
    row_bg=LIGHT if ri%2==0 else WHITE
    md=md_by_q.get(q_id,{}); pg=pg_by_q.get(q_id,{})
    md_vgs=md.get('vgs_yes'); pg_vgs=pg.get('vgs')
    md_sig=md.get('signal',''); pg_sig=pg.get('signal','')
    vals=[q_id, md.get('q_type',''), md.get('question',''),
          md_vgs, md_sig, pg_vgs, pg_sig, '—','—','—']
    for ci,v in enumerate(vals,1):
        bg=row_bg; fg='000000'; bold=False
        if ci==4 and v is not None: bg=sig_color(v); fg=WHITE; bold=True
        elif ci==5: bg=sig_color(md.get('vgs_yes')); fg=WHITE; bold=True
        elif ci==6 and v is not None: bg=sig_color(v); fg=WHITE; bold=True
        elif ci==7: bg=sig_color(pg.get('vgs')); fg=WHITE; bold=True
        elif ci in (8,9,10): bg=GREY; fg='999999'
        cell(ws10,ri,ci,v,bg=bg,fg=fg,bold=bold,align='left' if ci==3 else 'center')
    ws10.row_dimensions[ri].height=30

# ══════════════════════════════════════════════════════════════════
# SHEET: CrossModel_Bias
# ══════════════════════════════════════════════════════════════════
ws11 = wb.create_sheet('CrossModel_Bias')
title_row(ws11, 'Cross-Model Affirmation Bias | GT=yes acc vs GT=no acc | 2/5 models complete', 10)
ws11.freeze_panes = 'A3'
h2=['Q_ID','Question','MD_yes%','MD_no%','MD_gap%','PG_yes%','PG_no%','PG_gap%','Bias_direction_MD','Bias_direction_PG']
w2=[12,44,9,9,9,9,9,9,16,16]
set_cols(ws11,w2)
for i,(hh,ww) in enumerate(zip(h2,w2),1): hdr(ws11,2,i,hh,bg=TEAL)

for ri,q_id in enumerate(all_qids,3):
    row_bg=LIGHT if ri%2==0 else WHITE
    md=md_by_q.get(q_id,{}); pg=pg_by_q.get(q_id,{})
    md_gap=md.get('bias_gap'); pg_gap=pg.get('bias_gap')
    md_dir='Affirmation' if md_gap and md_gap>15 else ('Negation' if md_gap and md_gap<-15 else 'Balanced')
    pg_dir='Affirmation' if pg_gap and pg_gap>15 else ('Negation' if pg_gap and pg_gap<-15 else 'Balanced')
    vals=[q_id,md.get('question',''),
          md.get('yes_acc'),md.get('no_acc'),md_gap,
          pg.get('yes_acc'),pg.get('no_acc'),pg_gap,
          md_dir,pg_dir]
    for ci,v in enumerate(vals,1):
        bg=row_bg; fg='000000'; bold=False
        if ci in (5,8) and v is not None:
            bg=RED if v>25 else (AMBER if v>10 else (GREEN if abs(v)<=10 else BLUE))
            fg=WHITE if abs(v)>10 else '000000'; bold=True
        elif ci in (9,10):
            bg=RED if v=='Affirmation' else (BLUE if v=='Negation' else GREEN)
            fg=WHITE; bold=True
        cell(ws11,ri,ci,v,bg=bg,fg=fg,bold=bold,align='left' if ci==2 else 'center')
    ws11.row_dimensions[ri].height=30

# ══════════════════════════════════════════════════════════════════
# SHEET: LangPrior_Analysis
# ══════════════════════════════════════════════════════════════════
ws12 = wb.create_sheet('LangPrior_Analysis')
title_row(ws12, 'Language Prior Case Study | When does model answer with vs without image?', 11)
ws12.freeze_panes = 'A3'
h3=['Q_ID','Question','MD_ModeA_ans','MD_ModeC_ans','MD_case',
    'PG_ModeA_ans','PG_ModeC_ans','PG_case','MD_VGS','PG_VGS','Finding']
w3=[12,40,13,13,14,13,13,14,9,9,30]
set_cols(ws12,w3)
for i,(hh,ww) in enumerate(zip(h3,w3),1): hdr(ws12,2,i,hh,bg=TEAL)

# Build answer distributions per question per model
def ans_dist(rows, q_id):
    qr=[r for r in rows if r['q_id']==q_id]
    if not qr: return 'N/A'
    c=Counter(r['extracted_ans'] for r in qr)
    total=len(qr)
    parts=[]
    for k in ['yes','no','unclear']:
        if c[k]>0: parts.append('%s:%d%%'%(k,round(c[k]/total*100)))
    return ' '.join(parts)

md_main_all=[r for r in md_A if r['use_for']=='main_eval']
md_C_main  =[r for r in md_C if r['use_for']=='main_eval']
pg_main_all=[r for r in pg_A if r['use_for']=='main_eval']
pg_C_main  =[r for r in pg_C if r['use_for']=='main_eval']

case_colors={'Prior dominant':RED,'Visual grounding':GREEN,'Refuses without image':BLUE,'Mixed':AMBER}

for ri,q_id in enumerate(all_qids,3):
    row_bg=LIGHT if ri%2==0 else WHITE
    md=md_by_q.get(q_id,{}); pg=pg_by_q.get(q_id,{})

    md_a=ans_dist(md_main_all,q_id)
    md_c=ans_dist(md_C_main,q_id)
    pg_a=ans_dist(pg_main_all,q_id)
    pg_c=ans_dist(pg_C_main,q_id)

    # Classify case for each model
    md_vgs=md.get('vgs_yes',0) or 0
    pg_vgs=pg.get('vgs',0) or 0

    def classify(vgs, c_dist):
        if 'unclear:' in c_dist and int(c_dist.split('unclear:')[1].split('%')[0])>50:
            return 'Refuses without image'
        if abs(vgs)<5: return 'Prior dominant'
        if vgs>=20: return 'Visual grounding'
        return 'Mixed'

    md_case=classify(md_vgs, md_c)
    pg_case=classify(pg_vgs, pg_c)

    # Finding
    finding=''
    if md_case=='Prior dominant' and pg_case=='Visual grounding':
        finding='PG uses image, MD does not'
    elif md_case=='Visual grounding' and pg_case=='Prior dominant':
        finding='MD uses image, PG does not'
    elif md_case=='Prior dominant' and pg_case=='Prior dominant':
        finding='Both answer from prior — language prior confirmed'
    elif md_case=='Visual grounding' and pg_case=='Visual grounding':
        finding='Both visually grounded'
    elif pg_case=='Refuses without image':
        finding='PG refuses without image — needs visual input'
    else:
        finding='Mixed pattern'

    vals=[q_id,md.get('question',''),md_a,md_c,md_case,pg_a,pg_c,pg_case,
          md.get('vgs_yes'),pg.get('vgs'),finding]
    for ci,v in enumerate(vals,1):
        bg=row_bg; fg='000000'; bold=False
        if ci==5:
            bg=case_colors.get(v,GREY); fg=WHITE; bold=True
        elif ci==8:
            bg=case_colors.get(v,GREY); fg=WHITE; bold=True
        elif ci in (9,10) and v is not None:
            bg=sig_color(v); fg=WHITE; bold=True
        cell(ws12,ri,ci,v,bg=bg,fg=fg,bold=bold,align='left' if ci in (2,11) else 'center')
    ws12.row_dimensions[ri].height=35

# ══════════════════════════════════════════════════════════════════
# SHEET: Question_Tracker
# ══════════════════════════════════════════════════════════════════
ws13 = wb.create_sheet('Question_Tracker')
title_row(ws13, 'Question Status Tracker | Action Recommendations (updates as models complete)', 11)
ws13.freeze_panes = 'A3'
h4=['Q_ID','Type','Question','MD_signal','MD_VGS','PG_signal','PG_VGS',
    'SV_signal','LV_signal','IV_signal','Provisional_Action']
w4=[12,11,44,11,9,11,9,11,11,11,18]
set_cols(ws13,w4)
for i,(hh,ww) in enumerate(zip(h4,w4),1): hdr(ws13,2,i,hh,bg=TEAL)

for ri,q_id in enumerate(all_qids,3):
    row_bg=LIGHT if ri%2==0 else WHITE
    md=md_by_q.get(q_id,{}); pg=pg_by_q.get(q_id,{})
    # Provisional action based on 2 models
    md_sig=md.get('signal',''); pg_sig=pg.get('signal','')
    if md_sig=='NONE' and pg_sig=='NONE': prov='BASELINE ONLY'
    elif md_sig in ('STRONG','MODERATE') or pg_sig in ('STRONG','MODERATE'): prov='KEEP (verify)'
    elif md_sig=='WEAK' and pg_sig=='WEAK': prov='INVESTIGATE'
    else: prov='PENDING'
    vals=[q_id,md.get('q_type',''),md.get('question',''),
          md_sig,md.get('vgs_yes'),pg_sig,pg.get('vgs'),
          '—','—','—',prov]
    for ci,v in enumerate(vals,1):
        bg=row_bg; fg='000000'; bold=False
        if ci==4: bg=sig_color(md.get('vgs_yes')); fg=WHITE; bold=True
        elif ci==5 and v is not None: bg=sig_color(v); fg=WHITE; bold=True
        elif ci==6: bg=sig_color(pg.get('vgs')); fg=WHITE; bold=True
        elif ci==7 and v is not None: bg=sig_color(v); fg=WHITE; bold=True
        elif ci in (8,9,10): bg=GREY; fg='999999'
        elif ci==11:
            bg=GREEN if prov.startswith('KEEP') else (RED if prov=='BASELINE ONLY' else (AMBER if prov=='INVESTIGATE' else GREY))
            fg=WHITE if prov!='PENDING' else '000000'; bold=True
        cell(ws13,ri,ci,v,bg=bg,fg=fg,bold=bold,align='left' if ci==3 else 'center')
    ws13.row_dimensions[ri].height=32

OUT='/home2/muskan.singh/results/adverse_weather_results_v2.xlsx'
wb.save(OUT)
print('Saved: %s' % OUT)
print('Sheets: %s' % str(wb.sheetnames))
