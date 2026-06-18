import csv, json, openpyxl
from collections import defaultdict, Counter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Load data ─────────────────────────────────────────────────────
def load_csv(path):
    with open(path, newline='', encoding='utf-8') as f:
        return list(csv.DictReader(f))

def load_json(path):
    with open(path) as f:
        return json.load(f)

BASE = '/home2/muskan.singh/results'
md_A  = load_csv('%s/adverse_weather_moondream_v2.csv' % BASE)
md_C  = load_csv('%s/adverse_weather_moondream_noimage_v2.csv' % BASE)
pg_A  = load_csv('%s/adverse_weather_paligemma_v2.csv' % BASE)
pg_C  = load_csv('%s/adverse_weather_paligemma_noimage_v2.csv' % BASE)
sv_A  = load_csv('%s/adverse_weather_smolvlm_v2.csv' % BASE)
sv_C  = load_csv('%s/adverse_weather_smolvlm_noimage_v2.csv' % BASE)
md_an = load_json('%s/aw_moondream_v2_analysis.json' % BASE)
pg_an = load_json('%s/aw_paligemma_v2_analysis.json' % BASE)
sv_an = load_json('%s/aw_smolvlm_v2_analysis.json' % BASE)

manifest = load_json('/home2/muskan.singh/aw_manifest_v2.json')
var_lookup = {}
for entry in manifest:
    for q in entry['questions']:
        if q.get('use_for') == 'linguistic_variation':
            var_lookup[q['q_id']] = {
                'variation_type': q.get('variation_type', ''),
                'source_q_id': q.get('source_q_id', '')
            }

# ── Styles ────────────────────────────────────────────────────────
NAVY='0D1B2A'; TEAL='0E7C7B'; AMBER='E8A838'; RED='C0392B'
GREEN='27AE60'; LIGHT='EAF4FB'; WHITE='FFFFFF'; GREY='F2F2F2'
BLUE='2980B9'; PURPLE='7C3AED'

thin = Border(
    left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'))

def hdr(ws, row, col, val, bg=NAVY, fg=WHITE, wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill('solid', fgColor=bg)
    c.font = Font(color=fg, bold=True, name='Calibri', size=10)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    return c

def cell(ws, row, col, val, bg=WHITE, fg='000000', bold=False, align='center'):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill('solid', fgColor=bg)
    c.font = Font(color=fg, bold=bold, name='Calibri', size=10)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border = thin
    return c

def title_row(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.fill = PatternFill('solid', fgColor=NAVY)
    c.font = Font(color=WHITE, bold=True, name='Calibri', size=13)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

def set_cols(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def sig_color(v):
    if v is None: return GREY
    if v >= 50:   return GREEN
    if v >= 20:   return AMBER
    if v >= 5:    return 'E67E22'
    return RED

def acc_color(v):
    if v is None: return GREY
    if v >= 70:   return GREEN
    if v >= 50:   return AMBER
    return RED

action_colors = {'KEEP': GREEN, 'INVESTIGATE': AMBER, 'BASELINE ONLY': BLUE, 'REDESIGN': RED}

def get_action(s):
    if s['signal'] == 'NONE' and s.get('unclear', 0) / max(s.get('n', 1), 1) > 0.4:
        return 'REDESIGN'
    if s['signal'] == 'NONE': return 'BASELINE ONLY'
    if s['signal'] == 'WEAK': return 'INVESTIGATE'
    return 'KEEP'

# ── Helper: questions sheet ───────────────────────────────────────
def build_questions_sheet(ws, model_name, summary, vgs_note):
    title_row(ws, '%s | Per-Question Analysis | main_eval original images' % model_name, 17)
    ws.freeze_panes = 'A3'
    h = ['Q_ID','Type','Question','N','GT=yes','GT=no',
         'A_acc%','C_acc%','VGS%','VGS_yes%','VGS_no%',
         'yes_acc%','no_acc%','Bias_gap%','Unclear','Signal','Action']
    w = [12,12,44,6,7,6,8,8,10,10,10,9,9,10,8,10,15]
    set_cols(ws, w)
    for i,(hh,ww) in enumerate(zip(h,w),1): hdr(ws,2,i,hh,bg=TEAL)
    ws.row_dimensions[2].height = 22
    for ri, q in enumerate(sorted(summary, key=lambda x: x['q_id']), 3):
        action = get_action(q)
        row_bg = LIGHT if ri % 2 == 0 else WHITE
        vals = [q['q_id'],q['q_type'],q['question'],q['n'],q['gt_yes'],q['gt_no'],
                q['a_acc'],q['c_acc'],q['vgs'],q['vgs_yes'],q['vgs_no'],
                q['yes_acc'],q['no_acc'],q['bias_gap'],q['unclear'],q['signal'],action]
        for ci, v in enumerate(vals, 1):
            bg=row_bg; fg='000000'; bold=False
            if ci == 16: bg=sig_color(q['vgs_yes'] if q['vgs_yes'] is not None else q['vgs']); fg=WHITE; bold=True
            elif ci == 17: bg=action_colors.get(action, GREY); fg=WHITE; bold=True
            elif ci in (7,8):
                bg=acc_color(v) if v else row_bg
                fg=WHITE if v and (v >= 70 or v < 50) else '000000'
            elif ci in (9,10,11) and v is not None:
                bg=GREEN if v>40 else (AMBER if v>10 else (RED if v<-10 else row_bg))
                fg=WHITE if abs(v) > 10 else '000000'
            elif ci == 14 and v is not None:
                bg=RED if v>25 else (AMBER if v>10 else (GREEN if abs(v)<=10 else BLUE))
                fg=WHITE if abs(v) > 10 else '000000'; bold=True
            cell(ws, ri, ci, v, bg=bg, fg=fg, bold=bold, align='left' if ci==3 else 'center')
        ws.row_dimensions[ri].height = 32
    # Note row
    note_row = len(summary) + 4
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=17)
    nc = ws.cell(row=note_row, column=1, value=vgs_note)
    nc.fill = PatternFill('solid', fgColor='FFF3CD')
    nc.font = Font(name='Calibri', size=9, italic=True, color='7C4700')
    nc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[note_row].height = 30

# ── Helper: weather sheet ─────────────────────────────────────────
def build_weather_sheet(ws, model_name, weather_data, main_rows):
    title_row(ws, '%s | Accuracy by Weather Condition' % model_name, 9)
    ws.freeze_panes = 'A3'
    h = ['Condition','N_images','Overall_acc%','GT=yes_acc%','GT=no_acc%','Bias_gap%','GT_yes_n','GT_no_n','Unclear%']
    w = [13,10,13,13,13,11,11,10,10]
    set_cols(ws, w)
    for i,(hh,ww) in enumerate(zip(h,w),1): hdr(ws,2,i,hh,bg=TEAL)
    for ri, wd in enumerate(weather_data, 3):
        row_bg = LIGHT if ri % 2 == 0 else WHITE
        gap = round(wd['yes_acc']-wd['no_acc'],1) if wd.get('yes_acc') is not None and wd.get('no_acc') is not None else None
        unc_rows = [r for r in main_rows if r['weather']==wd['weather']]
        unc = round(sum(r['extracted_ans']=='unclear' for r in unc_rows)/len(unc_rows)*100,1) if unc_rows else 0
        vals = [wd['weather'].capitalize(),wd['n'],wd['acc'],wd.get('yes_acc'),wd.get('no_acc'),gap,wd['gt_yes'],wd['gt_no'],unc]
        for ci, v in enumerate(vals, 1):
            bg=row_bg; fg='000000'; bold=False
            if ci == 3: bg=acc_color(v); fg=WHITE; bold=True
            elif ci == 6 and v is not None:
                bg=RED if v>25 else (AMBER if v>10 else (BLUE if v<-25 else (AMBER if v<-10 else GREEN)))
                fg=WHITE; bold=True
            cell(ws, ri, ci, v, bg=bg, fg=fg, bold=bold)
        ws.row_dimensions[ri].height = 22
    # Overall
    all_yes=[r for r in main_rows if r['ground_truth']=='yes']
    all_no =[r for r in main_rows if r['ground_truth']=='no']
    ov=round(sum(r['correct']=='True' for r in main_rows)/len(main_rows)*100,1)
    ya=round(sum(r['correct']=='True' for r in all_yes)/len(all_yes)*100,1)
    na=round(sum(r['correct']=='True' for r in all_no)/len(all_no)*100,1)
    gap=round(ya-na,1)
    unc=round(sum(r['extracted_ans']=='unclear' for r in main_rows)/len(main_rows)*100,1)
    for ci,v in enumerate(['OVERALL',len(main_rows),ov,ya,na,gap,len(all_yes),len(all_no),unc],1):
        cell(ws,8,ci,v,bg=NAVY,fg=WHITE,bold=True)

# ── Helper: SCS sheet ─────────────────────────────────────────────
def build_scs_sheet(ws, model_name, scs_data):
    title_row(ws, '%s | Scene-Change Sensitivity (SCS) | Clear -> Fog/Rain/Snow' % model_name, 10)
    ws.freeze_panes = 'A3'
    h=['Q_ID','Question','N_pairs','SCS_overall%','SCS_fog%','SCS_rain%','SCS_snow%','GT_flip%','CDR%','Notes']
    w=[12,44,8,13,10,10,10,11,10,32]
    set_cols(ws, w)
    for i,(hh,ww) in enumerate(zip(h,w),1): hdr(ws,2,i,hh,bg=TEAL)
    for ri, d in enumerate(scs_data, 3):
        row_bg = LIGHT if ri % 2 == 0 else WHITE
        scs=d.get('scs'); gt_flip=d.get('gt_flip'); cdr=d.get('cdr')
        note=''
        if scs is not None:
            if scs>=50:   note='High — model responds to weather change'
            elif scs>=25: note='Moderate sensitivity'
            elif gt_flip is not None and gt_flip<10: note='Low GT flip — weather does not change GT here'
            else:         note='Low — model ignores weather change'
        vals=[d['q_id'],d.get('question',''),d.get('n_pairs'),scs,
              d.get('scs_fog'),d.get('scs_rain'),d.get('scs_snow'),gt_flip,cdr,note]
        for ci, v in enumerate(vals, 1):
            bg=row_bg; fg='000000'; bold=False
            if ci==4 and v is not None:
                bg=GREEN if v>=50 else (AMBER if v>=25 else RED); fg=WHITE; bold=True
            elif ci in (5,6,7) and v is not None:
                bg=GREEN if v>=50 else (AMBER if v>=25 else RED); fg=WHITE
            cell(ws, ri, ci, v, bg=bg, fg=fg, bold=bold, align='left' if ci in (2,10) else 'center')
        ws.row_dimensions[ri].height = 30

# ── Helper: linguistic sheet ──────────────────────────────────────
def build_ling_sheet(ws, model_name, ling_data):
    title_row(ws, '%s | Linguistic Variation Analysis | Paraphrase Consistency + Negation Accuracy' % model_name, 9)
    ws.freeze_panes = 'A3'
    h=['Source_Q_ID','Question','Para_pairs','Para_consistent%','Neg_pairs','Neg_acc%','Neg_flip%','Signal','Notes']
    w=[14,44,11,16,10,10,12,12,35]
    set_cols(ws, w)
    for i,(hh,ww) in enumerate(zip(h,w),1): hdr(ws,2,i,hh,bg=TEAL)
    sig_col={'ROBUST':GREEN,'MODERATE':AMBER,'INCONSISTENT':RED,'NEG FAILURE':RED,'N/A':GREY}
    for ri, d in enumerate(ling_data, 3):
        row_bg = LIGHT if ri % 2 == 0 else WHITE
        pp=d.get('para_pct'); na=d.get('neg_acc')
        note=''
        if pp is not None and pp<60:      note='Unstable — changes across paraphrases'
        elif na is not None and na<40:    note='Negation failure — model does not flip correctly'
        elif pp is not None and pp>=80 and na is not None and na>=60:
            note='Robust — consistent paraphrase + handles negation'
        vals=[d['source_q_id'],d.get('question',''),d.get('para_pairs'),pp,
              d.get('neg_pairs'),na,d.get('neg_flip'),d.get('signal'),note]
        for ci, v in enumerate(vals, 1):
            bg=row_bg; fg='000000'; bold=False
            if ci==4 and v is not None:
                bg=GREEN if v>=80 else (AMBER if v>=60 else RED); fg=WHITE; bold=True
            elif ci==6 and v is not None:
                bg=GREEN if v>=60 else (AMBER if v>=40 else RED); fg=WHITE; bold=True
            elif ci==8:
                bg=sig_col.get(v, GREY); fg=WHITE if v!='N/A' else '000000'; bold=True
            cell(ws, ri, ci, v, bg=bg, fg=fg, bold=bold, align='left' if ci in (2,9) else 'center')
        ws.row_dimensions[ri].height = 32

# ── Helper: Mode C segregation ────────────────────────────────────
def build_modec_seg_sheet(ws, models_data):
    
    n_models = len(models_data)
    n_cols = 2 + n_models * 4  # Q_ID + Question + (4 cols per model)
    title_row(ws, 'Mode C Segregation | Does the model refer to the image at all?', n_cols)
    ws.freeze_panes = 'A3'

    # Build header
    base_h = ['Q_ID', 'Question']
    base_w = [12, 40]
    model_cols = []
    for mname, _, _, _ in models_data:
        short = mname.split()[0]  # Moondream, PaliGemma, SmolVLM
        model_cols += ['%s_refuses%%' % short, '%s_prior%%' % short,
                       '%s_visual%%' % short, '%s_case' % short]
        base_w += [11, 11, 11, 14]

    all_h = base_h + model_cols
    set_cols(ws, base_w)
    for i, hh in enumerate(all_h, 1):
        hdr(ws, 2, i, hh, bg=TEAL)
    ws.row_dimensions[2].height = 22

    # Get all q_ids from first model
    first_A = models_data[0][1]
    all_qids = sorted(set(r['q_id'] for r in first_A
                         if r['use_for']=='main_eval' and r['is_augmented']=='False'))

    case_colors = {
        'Prior dominant': RED,
        'Visual influence': GREEN,
        'Refuses w/o image': BLUE,
        'Mixed': AMBER
    }

    for ri, q_id in enumerate(all_qids, 3):
        row_bg = LIGHT if ri % 2 == 0 else WHITE
        # Get question text
        q_txt = next((r['question'] for r in first_A if r['q_id']==q_id), '')
        row_vals = [q_id, q_txt]

        for mname, mA, mC, vgs_key in models_data:
            # Build per-image lookup for Mode C
            c_idx = {(r['image_id'], r['q_id']): r for r in mC}

            # Get main_eval original pairs for this question
            a_rows = [r for r in mA if r['q_id']==q_id
                      and r['use_for']=='main_eval' and r['is_augmented']=='False']

            if not a_rows:
                row_vals += [None, None, None, 'N/A']
                continue

            refuses = 0   # Mode C unclear
            same    = 0   # Mode A == Mode C (prior)
            diff    = 0   # Mode A != Mode C (visual)
            total   = 0

            for r in a_rows:
                c = c_idx.get((r['image_id'], r['q_id']))
                if not c: continue
                total += 1
                if c['extracted_ans'] == 'unclear':
                    refuses += 1
                elif r['extracted_ans'] == c['extracted_ans']:
                    same += 1
                else:
                    diff += 1

            if total == 0:
                row_vals += [None, None, None, 'N/A']
                continue

            ref_pct  = round(refuses/total*100, 1)
            same_pct = round(same/total*100, 1)
            diff_pct = round(diff/total*100, 1)

            # Classify
            if ref_pct >= 50:
                case = 'Refuses w/o image'
            elif same_pct >= 70:
                case = 'Prior dominant'
            elif diff_pct >= 50:
                case = 'Visual influence'
            else:
                case = 'Mixed'

            row_vals += [ref_pct, same_pct, diff_pct, case]

        # Write row
        for ci, v in enumerate(row_vals, 1):
            bg = row_bg; fg = '000000'; bold = False
            if ci == 1 or ci == 2:
                pass
            elif (ci - 2) % 4 == 0:  # case column
                bg = case_colors.get(v, GREY)
                fg = WHITE if v != 'N/A' else '000000'
                bold = True
            elif (ci - 2) % 4 == 1:  # refuses% — blue if high
                if v is not None and v >= 50: bg = BLUE; fg = WHITE
            elif (ci - 2) % 4 == 2:  # prior% — red if high
                if v is not None and v >= 70: bg = RED; fg = WHITE
            elif (ci - 2) % 4 == 3:  # visual% — green if high
                if v is not None and v >= 50: bg = GREEN; fg = WHITE
            cell(ws, ri, ci, v, bg=bg, fg=fg, bold=bold,
                 align='left' if ci == 2 else 'center')
        ws.row_dimensions[ri].height = 30

    # Add legend
    legend_row = len(all_qids) + 4
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=n_cols)
    lc = ws.cell(row=legend_row, column=1,
        value='CASE DEFINITIONS: "Refuses w/o image" = Mode C unclear ≥50% (model needs visual input) | '
              '"Prior dominant" = Mode A==Mode C answer ≥70% (language prior) | '
              '"Visual influence" = Mode A!=Mode C ≥50% (image changes response) | '
              '"Mixed" = no dominant pattern')
    lc.fill = PatternFill('solid', fgColor='EAF4FB')
    lc.font = Font(name='Calibri', size=9, italic=True, color='0D1B2A')
    lc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[legend_row].height = 35

# ══════════════════════════════════════════════════════════════════
# Build workbook
# ══════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

md_main = [r for r in md_A if r['use_for']=='main_eval' and r['is_augmented']=='False']
pg_main = [r for r in pg_A if r['use_for']=='main_eval' and r['is_augmented']=='False']
sv_main = [r for r in sv_A if r['use_for']=='main_eval' and r['is_augmented']=='False']

md_by_q = {s['q_id']: s for s in md_an['summary']}
pg_by_q = {s['q_id']: s for s in pg_an['summary']}
sv_by_q = {s['q_id']: s for s in sv_an['summary']}
all_qids = sorted(md_by_q.keys())

# ── README ────────────────────────────────────────────────────────
ws = wb.active; ws.title = 'README'
ws.column_dimensions['A'].width = 24
ws.column_dimensions['B'].width = 85
title_row(ws, 'Adverse Weather Results Workbook | Can VLMs Understand Driving? | IIIT Hyderabad', 2)
readme = [
    ('PROJECT',       'Can VLMs Understand Driving? | IIIT Hyderabad | Supervisor: Shankar Gangisetty'),
    ('CATEGORY',      'Adverse Weather | BDD100K 108 images + 72 counterfactual | Manifest v2 — 13,151 pairs'),
    ('MODELS DONE',   'Moondream 2B | PaliGemma 3B | SmolVLM 2.2B'),
    ('MODELS PENDING','LLaVA-OV 8B | InternVL3 8B'),
    ('',''),
    ('VGS (ours)',     'Mode A accuracy - Mode C accuracy. NOTE: Sir flagged this needs rethinking (IoU-like metric). Current VGS gives directional signal but is coarse.'),
    ('VGS_yes',       'VGS on GT=yes rows only. Clean signal for confounded models (Moondream/SmolVLM default to No on blank image).'),
    ('Mode C note MD','Moondream: ~100% No on blank images — visual default. Use VGS_yes only. VGS_overall is confounded.'),
    ('Mode C note SV','SmolVLM: 100% No on blank images — same confound as Moondream. Use VGS_yes only.'),
    ('Mode C note PG','PaliGemma: says unanswerable (kept as unclear=incorrect). VGS_overall valid — unanswerable is uniform across GT values.'),
    ('SCS',           'Scene-Change Sensitivity = % of clear->augmented pairs where model answer flipped.'),
    ('CDR',           'Correct Direction Rate = % of flipped pairs where flip was in correct direction.'),
    ('Para_consistent','% of paraphrase variants where model gives same answer as original on same image.'),
    ('Neg_acc',       'Accuracy on negated variants. Low with high Neg_flip = negation failure.'),
    ('Bias_gap',      'GT=yes acc - GT=no acc. Positive = affirmation bias. Negative = reverse/negation bias.'),
    ('',''),
    ('Mode C Segregation','NEW: Per question, per model — % pairs where model refuses (unclear), same answer as Mode A (prior), or different answer (visual influence).'),
    ('',''),
    ('SHEETS',''),
    ('MD_v2_Questions','Moondream per-question: VGS, bias, signal, action'),
    ('MD_v2_Weather','Moondream accuracy by weather condition'),
    ('MD_v2_SCS','Moondream scene-change sensitivity on counterfactual pairs'),
    ('MD_v2_Linguistic','Moondream paraphrase consistency + negation accuracy'),
    ('PG_v2_Questions','PaliGemma per-question analysis'),
    ('PG_v2_Weather','PaliGemma weather breakdown'),
    ('PG_v2_SCS','PaliGemma SCS'),
    ('PG_v2_Linguistic','PaliGemma linguistic'),
    ('SV_v2_Questions','SmolVLM per-question analysis'),
    ('SV_v2_Weather','SmolVLM weather breakdown'),
    ('SV_v2_SCS','SmolVLM SCS'),
    ('SV_v2_Linguistic','SmolVLM linguistic'),
    ('Mode_C_Segregation','Key sheet: when does each model actually refer to image vs answer from prior?'),
    ('CrossModel_VGS','VGS comparison across all models (fills as models complete)'),
    ('CrossModel_Bias','Affirmation/negation bias gap comparison'),
    ('LangPrior_Analysis','Answer distribution Mode A vs Mode C per model'),
    ('Question_Tracker','Provisional action per question across all models'),
]
for ri, (label, text) in enumerate(readme, 2):
    bg = LIGHT if ri%2==0 else WHITE
    lc = ws.cell(row=ri, column=1, value=label)
    lc.fill = PatternFill('solid', fgColor=TEAL if label and label not in ('','SHEETS') else (AMBER if label=='SHEETS' else WHITE))
    lc.font = Font(color=WHITE if label else '000000', bold=bool(label), name='Calibri', size=10)
    lc.alignment = Alignment(horizontal='left', vertical='center')
    tc = ws.cell(row=ri, column=2, value=text)
    tc.fill = PatternFill('solid', fgColor=bg)
    tc.font = Font(name='Calibri', size=10)
    tc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[ri].height = 20

# ── Model sheets ──────────────────────────────────────────────────
ws2 = wb.create_sheet('MD_v2_Questions')
build_questions_sheet(ws2, 'Moondream 2B v2', md_an['summary'],
    'VGS note: Moondream defaults ~100% No on blank images. Use VGS_yes as clean signal, not VGS_overall.')

ws3 = wb.create_sheet('MD_v2_Weather')
build_weather_sheet(ws3, 'Moondream 2B v2', md_an['weather'], md_main)

ws4 = wb.create_sheet('MD_v2_SCS')
build_scs_sheet(ws4, 'Moondream 2B v2', md_an.get('scs', []))

ws5 = wb.create_sheet('MD_v2_Linguistic')
build_ling_sheet(ws5, 'Moondream 2B v2', md_an.get('linguistic', []))

ws6 = wb.create_sheet('PG_v2_Questions')
build_questions_sheet(ws6, 'PaliGemma 3B v2', pg_an['summary'],
    'VGS note: PaliGemma says unanswerable on blank (kept as unclear=incorrect). VGS_overall valid — unanswerable uniform across GT values.')

ws7 = wb.create_sheet('PG_v2_Weather')
build_weather_sheet(ws7, 'PaliGemma 3B v2', pg_an['weather'], pg_main)

ws8 = wb.create_sheet('PG_v2_SCS')
build_scs_sheet(ws8, 'PaliGemma 3B v2', pg_an['scs'])

ws9 = wb.create_sheet('PG_v2_Linguistic')
build_ling_sheet(ws9, 'PaliGemma 3B v2', pg_an['linguistic'])

ws10 = wb.create_sheet('SV_v2_Questions')
build_questions_sheet(ws10, 'SmolVLM 2.2B v2', sv_an['summary'],
    'VGS note: SmolVLM defaults 100% No on blank images (same confound as Moondream). Use VGS_yes. Model has REVERSE bias — GT=no acc 85.8% vs GT=yes acc 43.8%.')

ws11 = wb.create_sheet('SV_v2_Weather')
build_weather_sheet(ws11, 'SmolVLM 2.2B v2', sv_an['weather'], sv_main)

ws12 = wb.create_sheet('SV_v2_SCS')
build_scs_sheet(ws12, 'SmolVLM 2.2B v2', sv_an['scs'])

ws13 = wb.create_sheet('SV_v2_Linguistic')
build_ling_sheet(ws13, 'SmolVLM 2.2B v2', sv_an['linguistic'])

# ── Mode C Segregation ────────────────────────────────────────────
ws14 = wb.create_sheet('Mode_C_Segregation')
models_data = [
    ('Moondream 2B', md_A, md_C, 'vgs_yes'),
    ('PaliGemma 3B', pg_A, pg_C, 'vgs'),
    ('SmolVLM 2.2B', sv_A, sv_C, 'vgs_yes'),
]
build_modec_seg_sheet(ws14, models_data)

# ── CrossModel_VGS ────────────────────────────────────────────────
ws15 = wb.create_sheet('CrossModel_VGS')
title_row(ws15, 'Cross-Model VGS Comparison | MD/SV use VGS_yes | PG uses VGS_overall | 3/5 models complete', 12)
ws15.freeze_panes = 'A3'
h = ['Q_ID','Type','Question','MD_VGS_yes','MD_sig','PG_VGS','PG_sig','SV_VGS_yes','SV_sig','LV_VGS','IV_VGS','Consensus_so_far']
w = [12,11,40,11,10,10,10,11,10,10,10,14]
set_cols(ws15, w)
for i,(hh,ww) in enumerate(zip(h,w),1): hdr(ws15,2,i,hh,bg=TEAL)

for ri, q_id in enumerate(all_qids, 3):
    row_bg = LIGHT if ri%2==0 else WHITE
    md=md_by_q.get(q_id,{}); pg=pg_by_q.get(q_id,{}); sv=sv_by_q.get(q_id,{})
    md_v=md.get('vgs_yes'); pg_v=pg.get('vgs'); sv_v=sv.get('vgs_yes')
    md_s=md.get('signal',''); pg_s=pg.get('signal',''); sv_s=sv.get('signal','')
    # Consensus
    sigs = [s for s in [md_s, pg_s, sv_s] if s]
    strong = sigs.count('STRONG'); none_c = sigs.count('NONE')
    if strong >= 2:   consensus = 'STRONG (2+)'
    elif strong == 1: consensus = 'MIXED'
    elif none_c >= 2: consensus = 'NONE (2+)'
    else:             consensus = 'WEAK/MOD'
    cons_color = GREEN if 'STRONG' in consensus else (RED if 'NONE' in consensus else AMBER)
    vals = [q_id,md.get('q_type',''),md.get('question',''),
            md_v,md_s,pg_v,pg_s,sv_v,sv_s,'—','—',consensus]
    for ci, v in enumerate(vals, 1):
        bg=row_bg; fg='000000'; bold=False
        if ci==4 and v is not None: bg=sig_color(v); fg=WHITE; bold=True
        elif ci==5: bg=sig_color(md_v); fg=WHITE; bold=True
        elif ci==6 and v is not None: bg=sig_color(v); fg=WHITE; bold=True
        elif ci==7: bg=sig_color(pg_v); fg=WHITE; bold=True
        elif ci==8 and v is not None: bg=sig_color(v); fg=WHITE; bold=True
        elif ci==9: bg=sig_color(sv_v); fg=WHITE; bold=True
        elif ci in (10,11): bg=GREY; fg='999999'
        elif ci==12: bg=cons_color; fg=WHITE; bold=True
        cell(ws15, ri, ci, v, bg=bg, fg=fg, bold=bold, align='left' if ci==3 else 'center')
    ws15.row_dimensions[ri].height = 30

# ── CrossModel_Bias ───────────────────────────────────────────────
ws16 = wb.create_sheet('CrossModel_Bias')
title_row(ws16, 'Cross-Model Affirmation/Negation Bias | GT=yes acc - GT=no acc | 3/5 models', 11)
ws16.freeze_panes = 'A3'
h2=['Q_ID','Question','MD_yes%','MD_no%','MD_gap%','PG_yes%','PG_no%','PG_gap%','SV_yes%','SV_no%','SV_gap%']
w2=[12,40,9,9,9,9,9,9,9,9,9]
set_cols(ws16, w2)
for i,(hh,ww) in enumerate(zip(h2,w2),1): hdr(ws16,2,i,hh,bg=TEAL)

for ri, q_id in enumerate(all_qids, 3):
    row_bg = LIGHT if ri%2==0 else WHITE
    md=md_by_q.get(q_id,{}); pg=pg_by_q.get(q_id,{}); sv=sv_by_q.get(q_id,{})
    vals=[q_id, md.get('question',''),
          md.get('yes_acc'),md.get('no_acc'),md.get('bias_gap'),
          pg.get('yes_acc'),pg.get('no_acc'),pg.get('bias_gap'),
          sv.get('yes_acc'),sv.get('no_acc'),sv.get('bias_gap')]
    for ci, v in enumerate(vals, 1):
        bg=row_bg; fg='000000'; bold=False
        if ci in (5,8,11) and v is not None:
            bg=RED if v>25 else (AMBER if v>10 else (GREEN if abs(v)<=10 else (AMBER if v>-25 else BLUE)))
            fg=WHITE if abs(v)>10 else '000000'; bold=True
        cell(ws16, ri, ci, v, bg=bg, fg=fg, bold=bold, align='left' if ci==2 else 'center')
    ws16.row_dimensions[ri].height = 30

# ── LangPrior_Analysis ────────────────────────────────────────────
ws17 = wb.create_sheet('LangPrior_Analysis')
title_row(ws17, 'Language Prior Case Study | Answer distributions Mode A vs Mode C per model', 12)
ws17.freeze_panes = 'A3'
h3=['Q_ID','Question','MD_A_dist','MD_C_dist','MD_case',
    'PG_A_dist','PG_C_dist','PG_case',
    'SV_A_dist','SV_C_dist','SV_case','Finding']
w3=[12,36,14,14,15,14,14,15,14,14,15,25]
set_cols(ws17, w3)
for i,(hh,ww) in enumerate(zip(h3,w3),1): hdr(ws17,2,i,hh,bg=TEAL)

case_colors={'Prior dominant':RED,'Visual influence':GREEN,'Refuses w/o image':BLUE,'Mixed':AMBER}

def ans_dist_str(rows, q_id):
    qr=[r for r in rows if r['q_id']==q_id]
    if not qr: return 'N/A'
    c=Counter(r['extracted_ans'] for r in qr); total=len(qr)
    parts=[]
    for k in ['yes','no','unclear']:
        if c[k]>0: parts.append('%s:%d%%'%(k,round(c[k]/total*100)))
    return ' '.join(parts)

def classify_case(vgs, c_dist_str):
    unclear_pct = 0
    for part in c_dist_str.split():
        if part.startswith('unclear:'):
            unclear_pct = int(part.replace('unclear:','').replace('%',''))
    if unclear_pct >= 50: return 'Refuses w/o image'
    if abs(vgs or 0) < 5: return 'Prior dominant'
    if (vgs or 0) >= 20:  return 'Visual influence'
    return 'Mixed'

md_C_main=[r for r in md_C if r['use_for']=='main_eval']
pg_C_main=[r for r in pg_C if r['use_for']=='main_eval']
sv_C_main=[r for r in sv_C if r['use_for']=='main_eval']
md_A_main=[r for r in md_A if r['use_for']=='main_eval']
pg_A_main=[r for r in pg_A if r['use_for']=='main_eval']
sv_A_main=[r for r in sv_A if r['use_for']=='main_eval']

for ri, q_id in enumerate(all_qids, 3):
    row_bg = LIGHT if ri%2==0 else WHITE
    md=md_by_q.get(q_id,{}); pg=pg_by_q.get(q_id,{}); sv=sv_by_q.get(q_id,{})
    md_a=ans_dist_str(md_A_main,q_id); md_c=ans_dist_str(md_C_main,q_id)
    pg_a=ans_dist_str(pg_A_main,q_id); pg_c=ans_dist_str(pg_C_main,q_id)
    sv_a=ans_dist_str(sv_A_main,q_id); sv_c=ans_dist_str(sv_C_main,q_id)
    md_case=classify_case(md.get('vgs_yes'), md_c)
    pg_case=classify_case(pg.get('vgs'), pg_c)
    sv_case=classify_case(sv.get('vgs_yes'), sv_c)
    cases=[md_case,pg_case,sv_case]
    if cases.count('Prior dominant')>=2: finding='Language prior confirmed (2+ models)'
    elif cases.count('Visual influence')>=2: finding='Visually grounded (2+ models)'
    elif cases.count('Refuses w/o image')>=1: finding='At least one model refuses w/o image'
    else: finding='Mixed — model-specific behaviour'
    vals=[q_id,md.get('question',''),md_a,md_c,md_case,pg_a,pg_c,pg_case,sv_a,sv_c,sv_case,finding]
    for ci, v in enumerate(vals, 1):
        bg=row_bg; fg='000000'; bold=False
        if ci in (5,8,11): bg=case_colors.get(v,GREY); fg=WHITE; bold=True
        cell(ws17, ri, ci, v, bg=bg, fg=fg, bold=bold, align='left' if ci in (2,12) else 'center')
    ws17.row_dimensions[ri].height = 35

# ── Question_Tracker ──────────────────────────────────────────────
ws18 = wb.create_sheet('Question_Tracker')
title_row(ws18, 'Question Status Tracker | Provisional Action based on 3/5 models', 12)
ws18.freeze_panes = 'A3'
h4=['Q_ID','Type','Question','MD_sig','MD_VGS','PG_sig','PG_VGS',
    'SV_sig','SV_VGS','LV_sig','IV_sig','Provisional_Action']
w4=[12,11,40,10,9,10,9,10,9,10,10,18]
set_cols(ws18, w4)
for i,(hh,ww) in enumerate(zip(h4,w4),1): hdr(ws18,2,i,hh,bg=TEAL)

for ri, q_id in enumerate(all_qids, 3):
    row_bg = LIGHT if ri%2==0 else WHITE
    md=md_by_q.get(q_id,{}); pg=pg_by_q.get(q_id,{}); sv=sv_by_q.get(q_id,{})
    md_s=md.get('signal',''); pg_s=pg.get('signal',''); sv_s=sv.get('signal','')
    sigs=[s for s in [md_s,pg_s,sv_s] if s]
    strong=sigs.count('STRONG'); none_c=sigs.count('NONE'); mod=sigs.count('MODERATE')
    if none_c==3: prov='BASELINE ONLY'
    elif none_c>=2 and strong==0: prov='LIKELY BASELINE'
    elif strong>=2: prov='KEEP'
    elif strong>=1 or mod>=2: prov='KEEP (verify)'
    else: prov='INVESTIGATE'
    vals=[q_id,md.get('q_type',''),md.get('question',''),
          md_s,md.get('vgs_yes'),pg_s,pg.get('vgs'),
          sv_s,sv.get('vgs_yes'),'—','—',prov]
    for ci, v in enumerate(vals, 1):
        bg=row_bg; fg='000000'; bold=False
        if ci in (4,6,8): bg=sig_color(md.get('vgs_yes') if ci==4 else (pg.get('vgs') if ci==6 else sv.get('vgs_yes'))); fg=WHITE; bold=True
        elif ci in (5,7,9) and v is not None: bg=sig_color(v); fg=WHITE; bold=True
        elif ci in (10,11): bg=GREY; fg='999999'
        elif ci==12:
            bg=GREEN if prov=='KEEP' else (TEAL if prov=='KEEP (verify)' else (RED if 'BASELINE' in prov else (AMBER if prov=='INVESTIGATE' else GREY)))
            fg=WHITE; bold=True
        cell(ws18, ri, ci, v, bg=bg, fg=fg, bold=bold, align='left' if ci==3 else 'center')
    ws18.row_dimensions[ri].height = 32

OUT = '/home2/muskan.singh/results/adverse_weather_results_v3.xlsx'
wb.save(OUT)
print('Saved: %s' % OUT)
print('Sheets (%d): %s' % (len(wb.sheetnames), str(wb.sheetnames)))
