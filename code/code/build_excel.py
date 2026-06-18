import csv, json, openpyxl
from collections import defaultdict, Counter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def load_csv(path):
    with open(path, newline='', encoding='utf-8') as f:
        return list(csv.DictReader(f))

modeA = load_csv('/home2/muskan.singh/results/adverse_weather_moondream_v2.csv')
modeC = load_csv('/home2/muskan.singh/results/adverse_weather_moondream_noimage_v2.csv')

with open('/home2/muskan.singh/aw_manifest_v2.json') as f:
    manifest = json.load(f)

var_lookup = {}
for entry in manifest:
    for q in entry['questions']:
        if q.get('use_for') == 'linguistic_variation':
            var_lookup[q['q_id']] = {
                'variation_type': q.get('variation_type',''),
                'source_q_id':    q.get('source_q_id','')
            }

c_idx       = {(r['image_id'], r['q_id']): r for r in modeC}
main_orig   = [r for r in modeA if r['use_for']=='main_eval'     and r['is_augmented']=='False']
cf_rows     = [r for r in modeA if r['is_augmented']=='True']
ling_rows   = [r for r in modeA if r['use_for']=='linguistic_variation']

NAVY='0D1B2A'; TEAL='0E7C7B'; AMBER='E8A838'; RED='C0392B'
GREEN='27AE60'; LIGHT='EAF4FB'; WHITE='FFFFFF'; GREY='F2F2F2'; BLUE='2980B9'

thin = Border(
    left=Side(style='thin',color='CCCCCC'), right=Side(style='thin',color='CCCCCC'),
    top=Side(style='thin',color='CCCCCC'),  bottom=Side(style='thin',color='CCCCCC'))

def hdr(ws, row, col, val, bg=NAVY, fg=WHITE, wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill('solid', fgColor=bg)
    c.font = Font(color=fg, bold=True, name='Calibri', size=10)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    return c

def cell(ws, row, col, val, bg=WHITE, fg='000000', bold=False, align='center'):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill('solid', fgColor=bg)
    c.font = Font(color=fg, bold=bold, name='Calibri', size=10)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border = thin
    return c

def title_row(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.fill = PatternFill('solid', fgColor=NAVY)
    c.font = Font(color=WHITE, bold=True, name='Calibri', size=13)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

def set_cols(ws, widths):
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w

def sig_color(v):
    if v is None: return GREY
    if v >= 50:   return GREEN
    if v >= 20:   return AMBER
    if v >= 5:    return 'E67E22'
    return RED

def acc_color(v):
    if v is None: return GREY
    if v >= 70:   return GREEN
    if v >= 50:   return AMBER
    return RED

def compute_q_stats(a_rows, c_index):
    by_q = defaultdict(list)
    for r in a_rows: by_q[r['q_id']].append(r)
    stats = {}
    for q_id, rows in by_q.items():
        n      = len(rows)
        a_acc  = sum(r['correct']=='True' for r in rows)/n
        c_rows = [c_index.get((r['image_id'],r['q_id'])) for r in rows]
        c_rows = [c for c in c_rows if c]
        c_acc  = sum(r['correct']=='True' for r in c_rows)/len(c_rows) if c_rows else None
        vgs    = round((a_acc - c_acc)*100,1) if c_acc is not None else None

        yes_A = [r for r in rows if r['ground_truth']=='yes']
        no_A  = [r for r in rows if r['ground_truth']=='no']
        yes_C = [c_index.get((r['image_id'],r['q_id'])) for r in yes_A]
        no_C  = [c_index.get((r['image_id'],r['q_id'])) for r in no_A]
        yes_C = [c for c in yes_C if c]
        no_C  = [c for c in no_C if c]

        vgs_yes = round((sum(r['correct']=='True' for r in yes_A)/len(yes_A) -
                         sum(r['correct']=='True' for r in yes_C)/len(yes_C))*100,1) \
                  if yes_A and yes_C else None
        vgs_no  = round((sum(r['correct']=='True' for r in no_A)/len(no_A) -
                         sum(r['correct']=='True' for r in no_C)/len(no_C))*100,1) \
                  if no_A and no_C else None

        yes_acc = round(sum(r['correct']=='True' for r in yes_A)/len(yes_A)*100,1) if yes_A else None
        no_acc  = round(sum(r['correct']=='True' for r in no_A)/len(no_A)*100,1)   if no_A  else None
        unclear = sum(r['extracted_ans']=='unclear' for r in rows)

        if vgs_yes is not None:
            sig = 'STRONG' if vgs_yes>=50 else ('MODERATE' if vgs_yes>=20 else ('WEAK' if vgs_yes>=5 else 'NONE'))
        else:
            sig = 'N/A'

        stats[q_id] = dict(
            q_id=q_id, q_type=rows[0]['q_type'], question=rows[0]['question'],
            n=n, gt_yes=len(yes_A), gt_no=len(no_A),
            a_acc=round(a_acc*100,1),
            c_acc=round(c_acc*100,1) if c_acc is not None else None,
            vgs=vgs, vgs_yes=vgs_yes, vgs_no=vgs_no,
            yes_acc=yes_acc, no_acc=no_acc,
            bias_gap=round(yes_acc-no_acc,1) if yes_acc and no_acc else None,
            unclear=unclear, signal=sig
        )
    return stats

q_stats = compute_q_stats(main_orig, c_idx)

action_colors = {'KEEP':GREEN,'INVESTIGATE':AMBER,'BASELINE ONLY':BLUE,'REDESIGN':RED}

def get_action(s):
    if s['signal']=='NONE' and s['unclear']/max(s['n'],1)>0.4: return 'REDESIGN'
    if s['signal']=='NONE': return 'BASELINE ONLY'
    if s['signal']=='WEAK': return 'INVESTIGATE'
    return 'KEEP'

wb = openpyxl.Workbook()

# ── SHEET 1: README ──────────────────────────────────────────────
ws = wb.active; ws.title = 'README'
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 85
title_row(ws, 'Adverse Weather Results Workbook | Can VLMs Understand Driving? | IIIT Hyderabad', 2)
readme = [
    ('PROJECT',         'Can VLMs Understand Driving? | IIIT Hyderabad | Supervisor: Shankar Gangisetty'),
    ('CATEGORY',        'Adverse Weather | BDD100K 108 images (rainy/foggy/snowy/night/clear) + 72 counterfactual'),
    ('MANIFEST v2',     '13,151 QA pairs: 2,859 original + 1,979 counterfactual + 8,313 linguistic variations'),
    ('MODELS DONE',     'Moondream 2B'),
    ('MODELS PENDING',  'PaliGemma 3B | SmolVLM 2.2B | LLaVA-OV 8B | InternVL3 8B'),
    ('',                ''),
    ('VGS',             'Visual Grounding Score = Mode A acc - Mode C acc. How much does the image contribute vs language prior?'),
    ('VGS_yes',         'VGS on GT=yes rows only. Clean signal for Moondream/SmolVLM (confounded blank-image behaviour).'),
    ('VGS_no',          'VGS on GT=no rows. Confounded for Moondream — blank image defaults to no. Do not interpret alone.'),
    ('SCS',             'Scene-Change Sensitivity = fraction of clear->augmented pairs where model answer flipped.'),
    ('CDR',             'Correct Direction Rate = fraction of flipped pairs where flip was in the correct direction.'),
    ('Para_consistent', 'Paraphrase consistency = model gives same answer to paraphrase variant as original question.'),
    ('Neg_acc',         'Negated variant accuracy = model correctly answers negated version of question.'),
    ('Bias_gap',        'GT=yes acc - GT=no acc. Positive = affirmation bias. Negative = negation/absence bias.'),
    ('Mode A',          'Standard inference with real image.'),
    ('Mode C',          'No-image baseline — blank black image. Measures language prior strength.'),
    ('',                ''),
    ('SHEETS',          ''),
    ('MD_v2_Questions', 'Per-question: accuracy, VGS, bias gap, signal, recommended action'),
    ('MD_v2_Weather',   'Accuracy by weather condition with bias breakdown'),
    ('MD_v2_Counterfactual', 'SCS per question — did model flip answer when weather changed?'),
    ('MD_v2_Linguistic','Paraphrase consistency + negated variant accuracy per question'),
    ('CrossModel_VGS',  'VGS_yes per question x model — fills as models complete'),
    ('Question_Tracker','Final action per question — KEEP / INVESTIGATE / BASELINE ONLY / REDESIGN'),
]
for ri,(label,text) in enumerate(readme,2):
    bg = LIGHT if ri%2==0 else WHITE
    lc = ws.cell(row=ri,column=1,value=label)
    lc.fill = PatternFill('solid', fgColor=TEAL if label and label not in ('','SHEETS') else (AMBER if label=='SHEETS' else WHITE))
    lc.font = Font(color=WHITE if label else '000000', bold=bool(label), name='Calibri', size=10)
    lc.alignment = Alignment(horizontal='left', vertical='center')
    tc = ws.cell(row=ri,column=2,value=text)
    tc.fill = PatternFill('solid', fgColor=bg)
    tc.font = Font(name='Calibri', size=10)
    tc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[ri].height = 20

# ── SHEET 2: MD_v2_Questions ─────────────────────────────────────
ws2 = wb.create_sheet('MD_v2_Questions')
title_row(ws2, 'Moondream 2B v2 | Per-Question Analysis | main_eval original images only', 17)
ws2.freeze_panes = 'A3'
h2 = ['Q_ID','Type','Question','N','GT=yes','GT=no',
      'A_acc%','C_acc%','VGS_overall%','VGS_yes%','VGS_no%',
      'yes_acc%','no_acc%','Bias_gap%','Unclear','Signal','Action']
w2 = [12,12,44,6,7,6,8,8,12,10,10,9,9,10,8,10,15]
set_cols(ws2,w2)
for i,(h,w) in enumerate(zip(h2,w2),1): hdr(ws2,2,i,h,bg=TEAL)
ws2.row_dimensions[2].height = 22

for ri,q in enumerate(sorted(q_stats.values(),key=lambda x:x['q_id']),3):
    action = get_action(q)
    row_bg = LIGHT if ri%2==0 else WHITE
    vals = [q['q_id'],q['q_type'],q['question'],q['n'],q['gt_yes'],q['gt_no'],
            q['a_acc'],q['c_acc'],q['vgs'],q['vgs_yes'],q['vgs_no'],
            q['yes_acc'],q['no_acc'],q['bias_gap'],q['unclear'],q['signal'],action]
    for ci,v in enumerate(vals,1):
        bg=row_bg; fg='000000'; bold=False
        if ci==16: bg=sig_color(q['vgs_yes']); fg=WHITE; bold=True
        elif ci==17: bg=action_colors.get(action,GREY); fg=WHITE; bold=True
        elif ci in (7,8):
            bg=acc_color(v) if v else row_bg
            fg=WHITE if v and (v>=70 or v<50) else '000000'
        elif ci in (9,10,11) and v is not None:
            bg=GREEN if v>40 else (AMBER if v>10 else (RED if v<-10 else row_bg))
            fg=WHITE if abs(v)>10 else '000000'
        elif ci==14 and v is not None:
            bg=RED if v>25 else (AMBER if v>10 else (GREEN if abs(v)<=10 else BLUE))
            fg=WHITE if abs(v)>10 else '000000'; bold=True
        cell(ws2,ri,ci,v,bg=bg,fg=fg,bold=bold,align='left' if ci==3 else 'center')
    ws2.row_dimensions[ri].height = 32

# ── SHEET 3: MD_v2_Weather ───────────────────────────────────────
ws3 = wb.create_sheet('MD_v2_Weather')
title_row(ws3, 'Moondream 2B v2 | Accuracy by Weather Condition', 9)
ws3.freeze_panes = 'A3'
h3 = ['Condition','N_images','Overall_acc%','GT=yes_acc%','GT=no_acc%','Bias_gap%','GT_yes_n','GT_no_n','Unclear%']
w3 = [13,10,13,13,13,11,11,10,10]
set_cols(ws3,w3)
for i,(h,w) in enumerate(zip(h3,w3),1): hdr(ws3,2,i,h,bg=TEAL)

by_w = defaultdict(list)
for r in main_orig: by_w[r['weather']].append(r)
for ri,wc in enumerate(['clear','rainy','foggy','snowy','night'],3):
    rows = by_w.get(wc,[])
    if not rows: continue
    row_bg = LIGHT if ri%2==0 else WHITE
    yes_r = [r for r in rows if r['ground_truth']=='yes']
    no_r  = [r for r in rows if r['ground_truth']=='no']
    ov  = round(sum(r['correct']=='True' for r in rows)/len(rows)*100,1)
    ya  = round(sum(r['correct']=='True' for r in yes_r)/len(yes_r)*100,1) if yes_r else None
    na  = round(sum(r['correct']=='True' for r in no_r)/len(no_r)*100,1)   if no_r  else None
    gap = round(ya-na,1) if ya and na else None
    unc = round(sum(r['extracted_ans']=='unclear' for r in rows)/len(rows)*100,1)
    vals=[wc.capitalize(),len(rows),ov,ya,na,gap,len(yes_r),len(no_r),unc]
    for ci,v in enumerate(vals,1):
        bg=row_bg; fg='000000'; bold=False
        if ci==3: bg=acc_color(v); fg=WHITE; bold=True
        elif ci==6 and v is not None:
            bg=RED if v>25 else (AMBER if v>10 else GREEN); fg=WHITE; bold=True
        cell(ws3,ri,ci,v,bg=bg,fg=fg,bold=bold)
    ws3.row_dimensions[ri].height = 22

# Overall row
all_yes=[r for r in main_orig if r['ground_truth']=='yes']
all_no =[r for r in main_orig if r['ground_truth']=='no']
ov_all =round(sum(r['correct']=='True' for r in main_orig)/len(main_orig)*100,1)
ya_all =round(sum(r['correct']=='True' for r in all_yes)/len(all_yes)*100,1)
na_all =round(sum(r['correct']=='True' for r in all_no)/len(all_no)*100,1)
gap_all=round(ya_all-na_all,1)
unc_all=round(sum(r['extracted_ans']=='unclear' for r in main_orig)/len(main_orig)*100,1)
for ci,v in enumerate(['OVERALL',len(main_orig),ov_all,ya_all,na_all,gap_all,len(all_yes),len(all_no),unc_all],1):
    cell(ws3,8,ci,v,bg=NAVY,fg=WHITE,bold=True)

# ── SHEET 4: MD_v2_Counterfactual ───────────────────────────────
ws4 = wb.create_sheet('MD_v2_Counterfactual')
title_row(ws4, 'Moondream 2B v2 | Scene-Change Sensitivity (SCS) | Clear -> Fog / Rain / Snow', 10)
ws4.freeze_panes = 'A3'
h4 = ['Q_ID','Question','N_pairs','SCS_overall%','SCS_fog%','SCS_rain%','SCS_snow%','GT_flip%','CDR%','Notes']
w4 = [12,44,8,13,10,10,10,11,10,32]
set_cols(ws4,w4)
for i,(h,w) in enumerate(zip(h4,w4),1): hdr(ws4,2,i,h,bg=TEAL)

orig_idx = {(r['image_id'],r['q_id']): r for r in main_orig}
cf_by_q  = defaultdict(list)
for r in cf_rows: cf_by_q[r['q_id']].append(r)
cf_q_ids = sorted(cf_by_q.keys())

for ri,q_id in enumerate(cf_q_ids,3):
    cf = cf_by_q[q_id]
    row_bg = LIGHT if ri%2==0 else WHITE
    pairs=0; flipped=0; gt_flips=0; cdr=0
    by_aug = defaultdict(lambda:[0,0])
    for r in cf:
        src = orig_idx.get((r['source_image_id'],q_id))
        if not src: continue
        pairs+=1
        aug_type=r['weather']
        by_aug[aug_type][0]+=1
        did_flip = r['extracted_ans']!=src['extracted_ans']
        if did_flip:
            flipped+=1; by_aug[aug_type][1]+=1
            if r['correct']=='True' and src['correct']=='False': cdr+=1
        if r['ground_truth']!=src['ground_truth']: gt_flips+=1
    scs      = round(flipped/pairs*100,1)      if pairs   else None
    scs_fog  = round(by_aug['foggy'][1]/by_aug['foggy'][0]*100,1)  if by_aug['foggy'][0]  else None
    scs_rain = round(by_aug['rainy'][1]/by_aug['rainy'][0]*100,1)  if by_aug['rainy'][0]  else None
    scs_snow = round(by_aug['snowy'][1]/by_aug['snowy'][0]*100,1)  if by_aug['snowy'][0]  else None
    gt_flip  = round(gt_flips/pairs*100,1)     if pairs   else None
    cdr_pct  = round(cdr/flipped*100,1)        if flipped else None
    note=''
    if scs is not None:
        if scs>=50:   note='High — model responds to weather change'
        elif scs>=25: note='Moderate sensitivity'
        elif gt_flip is not None and gt_flip<10: note='Low GT flip rate — weather does not change GT here'
        else:         note='Low — model ignores weather change'
    q_txt = cf[0]['question'] if cf else ''
    vals=[q_id,q_txt,pairs,scs,scs_fog,scs_rain,scs_snow,gt_flip,cdr_pct,note]
    for ci,v in enumerate(vals,1):
        bg=row_bg; fg='000000'; bold=False
        if ci==4 and v is not None:
            bg=GREEN if v>=50 else (AMBER if v>=25 else RED); fg=WHITE; bold=True
        elif ci in (5,6,7) and v is not None:
            bg=GREEN if v>=50 else (AMBER if v>=25 else RED); fg=WHITE
        cell(ws4,ri,ci,v,bg=bg,fg=fg,bold=bold,align='left' if ci in (2,10) else 'center')
    ws4.row_dimensions[ri].height = 30

# ── SHEET 5: MD_v2_Linguistic ────────────────────────────────────
ws5 = wb.create_sheet('MD_v2_Linguistic')
title_row(ws5, 'Moondream 2B v2 | Linguistic Variation Analysis | Paraphrase Consistency + Negation Accuracy', 9)
ws5.freeze_panes = 'A3'
h5 = ['Source_Q_ID','Question','Para_pairs','Para_consistent%','Neg_pairs','Neg_acc%','Neg_flip%','Signal','Notes']
w5 = [14,44,11,16,10,10,12,12,35]
set_cols(ws5,w5)
for i,(h,w) in enumerate(zip(h5,w5),1): hdr(ws5,2,i,h,bg=TEAL)

orig_main_idx = defaultdict(dict)
for r in main_orig: orig_main_idx[r['image_id']][r['q_id']] = r

para_by_src=defaultdict(list); neg_by_src=defaultdict(list)
for r in ling_rows:
    info  = var_lookup.get(r['q_id'],{})
    src_q = info.get('source_q_id','')
    vtype = info.get('variation_type','')
    if vtype=='paraphrase': para_by_src[src_q].append(r)
    elif vtype=='negated':  neg_by_src[src_q].append(r)

all_src_ids = sorted(set(list(para_by_src.keys())+list(neg_by_src.keys())))

for ri,src_q in enumerate(all_src_ids,3):
    row_bg=LIGHT if ri%2==0 else WHITE
    para=para_by_src.get(src_q,[]); neg=neg_by_src.get(src_q,[])
    para_pairs=0; para_ok=0
    for r in para:
        orig=orig_main_idx.get(r['image_id'],{}).get(src_q)
        if not orig: continue
        para_pairs+=1
        if r['extracted_ans']==orig['extracted_ans']: para_ok+=1
    neg_pairs=0; neg_correct=0; neg_flipped=0
    for r in neg:
        orig=orig_main_idx.get(r['image_id'],{}).get(src_q)
        if not orig: continue
        neg_pairs+=1
        if r['correct']=='True': neg_correct+=1
        if r['extracted_ans']!=orig['extracted_ans']: neg_flipped+=1
    para_pct = round(para_ok/para_pairs*100,1)     if para_pairs else None
    neg_acc  = round(neg_correct/neg_pairs*100,1)  if neg_pairs  else None
    neg_flip = round(neg_flipped/neg_pairs*100,1)  if neg_pairs  else None
    if para_pct is not None and neg_acc is not None:
        if   para_pct>=80 and neg_acc>=60: lsig='ROBUST'
        elif para_pct>=60 and neg_acc>=40: lsig='MODERATE'
        elif para_pct<60:                  lsig='INCONSISTENT'
        else:                              lsig='NEG FAILURE'
    else: lsig='N/A'
    sig_col={'ROBUST':GREEN,'MODERATE':AMBER,'INCONSISTENT':RED,'NEG FAILURE':RED,'N/A':GREY}
    note=''
    if para_pct is not None and para_pct<60:        note='Unstable — changes across paraphrases'
    elif neg_acc is not None and neg_acc<40:        note='Negation failure — model does not flip correctly'
    elif para_pct is not None and para_pct>=80 and neg_acc is not None and neg_acc>=60:
        note='Robust — consistent paraphrase + handles negation'
    src_q_txt=''
    for r in main_orig:
        if r['q_id']==src_q: src_q_txt=r['question']; break
    vals=[src_q,src_q_txt,para_pairs,para_pct,neg_pairs,neg_acc,neg_flip,lsig,note]
    for ci,v in enumerate(vals,1):
        bg=row_bg; fg='000000'; bold=False
        if ci==4 and v is not None:
            bg=GREEN if v>=80 else (AMBER if v>=60 else RED); fg=WHITE; bold=True
        elif ci==6 and v is not None:
            bg=GREEN if v>=60 else (AMBER if v>=40 else RED); fg=WHITE; bold=True
        elif ci==8:
            bg=sig_col.get(v,GREY); fg=WHITE if v!='N/A' else '000000'; bold=True
        cell(ws5,ri,ci,v,bg=bg,fg=fg,bold=bold,align='left' if ci in (2,9) else 'center')
    ws5.row_dimensions[ri].height = 32

# ── SHEET 6: CrossModel_VGS ──────────────────────────────────────
ws6 = wb.create_sheet('CrossModel_VGS')
title_row(ws6, 'Cross-Model VGS_yes Comparison | Adverse Weather (fills as models complete)', 9)
ws6.freeze_panes = 'A3'
models=['Moondream 2B','PaliGemma 3B','SmolVLM 2.2B','LLaVA-OV 8B','InternVL3 8B']
h6=['Q_ID','Question','MD_signal']+models+['Consensus']
w6=[12,44,11]+[13]*5+[12]
set_cols(ws6,w6)
for i,(h,w) in enumerate(zip(h6,w6),1): hdr(ws6,2,i,h,bg=TEAL)
for ri,q in enumerate(sorted(q_stats.values(),key=lambda x:x['q_id']),3):
    row_bg=LIGHT if ri%2==0 else WHITE
    cell(ws6,ri,1,q['q_id'],bg=row_bg)
    cell(ws6,ri,2,q['question'],bg=row_bg,align='left')
    cell(ws6,ri,3,q['signal'],bg=sig_color(q['vgs_yes']),fg=WHITE,bold=True)
    cell(ws6,ri,4,q['vgs_yes'],bg=sig_color(q['vgs_yes']),fg=WHITE,bold=True)
    for ci in range(5,9): cell(ws6,ri,ci,'—',bg=GREY,fg='999999')
    cell(ws6,ri,9,'PENDING',bg=GREY,fg='999999')
    ws6.row_dimensions[ri].height = 30

# ── SHEET 7: Question_Tracker ────────────────────────────────────
ws7 = wb.create_sheet('Question_Tracker')
title_row(ws7, 'Question Status Tracker | Action Recommendations (updates as models complete)', 10)
ws7.freeze_panes = 'A3'
h7=['Q_ID','Type','Question','MD_signal','MD_VGS_yes%','PG_signal','SV_signal','LV_signal','IV_signal','Final_Action']
w7=[12,12,44,11,12,11,11,11,11,16]
set_cols(ws7,w7)
for i,(h,w) in enumerate(zip(h7,w7),1): hdr(ws7,2,i,h,bg=TEAL)
for ri,q in enumerate(sorted(q_stats.values(),key=lambda x:x['q_id']),3):
    row_bg=LIGHT if ri%2==0 else WHITE
    action=get_action(q)
    vals=[q['q_id'],q['q_type'],q['question'],q['signal'],q['vgs_yes']]
    for ci,v in enumerate(vals,1):
        bg=row_bg; fg='000000'; bold=False
        if ci==4: bg=sig_color(q['vgs_yes']); fg=WHITE; bold=True
        elif ci==5 and v is not None: bg=sig_color(v); fg=WHITE; bold=True
        cell(ws7,ri,ci,v,bg=bg,fg=fg,bold=bold,align='left' if ci==3 else 'center')
    for ci in range(6,10): cell(ws7,ri,ci,'—',bg=GREY,fg='999999')
    cell(ws7,ri,10,action,bg=action_colors.get(action,GREY),fg=WHITE,bold=True)
    ws7.row_dimensions[ri].height = 32

OUT='/home2/muskan.singh/results/adverse_weather_results.xlsx'
wb.save(OUT)
print(f'Saved: {OUT}')
print(f'Sheets: {wb.sheetnames}')
